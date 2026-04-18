"""
5/8以上の361銘柄全員のデータを取得して chart_viewer.html に埋め込む
データソース:
  ①②③⑦    … irbank /results（10年）売上/営利率/EPS/自己資本比率
  ④⑧        … kabutan（5〜8年）営業CF/現金等残高
  ⑤⑥        … irbank /dividend（10年+）配当金、EPS と照合して配当性向
"""
import requests, re, time, sys, io, datetime, json, os, webbrowser
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE_DIR = Path(__file__).parent

HEADERS = {
    'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                   'AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36'),
    'Accept-Language': 'ja,en-US;q=0.9',
}
TREND_MIN_RATIO = 0.9
TODAY_YR = datetime.date.today().year

def parse_num(s):
    s = str(s).replace(',','').strip()
    s = re.sub(r'[^\d\.\-]','',s)
    try: return float(s) if s not in ('','-') else None
    except: return None

def parse_oku(s):
    """億円 → 百万円。億付きでなければ通常の parse_num"""
    s = str(s).strip().replace(',','')
    m = re.search(r'([\-\d\.]+)億', s)
    if m: return round(float(m.group(1)) * 100, 1)
    m2 = re.search(r'([\-\d\.]+)兆', s)
    if m2: return round(float(m2.group(1)) * 100_000, 1)
    return parse_num(s)

def yr_label(cell_text):
    """セルテキスト → "YYYY/MM" ラベル（1桁月も対応）"""
    m = re.search(r'(\d{4})[/年](\d{1,2})', cell_text)
    return f"{m.group(1)}/{int(m.group(2)):02d}" if m else None

# ── irbank メインページ取得（業種・決算発表日を一括取得） ──
def fetch_irbank_main(code):
    """IRBankの企業ページから業種と次回決算発表日を取得"""
    try:
        r = requests.get(f'https://irbank.net/{code}',
                         headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.content, 'html.parser')

        # 業種（パンくずから）
        sector = None
        for nav in soup.find_all(['nav', 'ol', 'ul']):
            anchors = nav.find_all('a')
            for i, a in enumerate(anchors):
                if 'ホーム' in a.get_text(strip=True):
                    if i + 1 < len(anchors):
                        sector = anchors[i + 1].get_text(strip=True)
                        break
            if sector:
                break

        # 次回決算発表日（/market/kessan?y=YYYY-MM-DD リンクから）
        earnings_date = None
        for a in soup.find_all('a', href=True):
            m = re.search(r'/market/kessan\?y=(\d{4}-\d{2}-\d{2})', a['href'])
            if m:
                earnings_date = m.group(1)
                break

        return sector, earnings_date
    except:
        return None, None

def fetch_sector(code):
    sector, _ = fetch_irbank_main(code)
    return sector

def check_trend(lst, min_ratio=None):
    if min_ratio is None:
        min_ratio = TREND_MIN_RATIO
    vals = [v for v in lst if v is not None]
    if len(vals) < 2: return None
    for i in range(1, len(vals)):
        prev, curr = vals[i-1], vals[i]
        if prev == 0:
            if curr <= 0: return False
            continue
        if curr < prev * min_ratio: return False
    return True

def no_cut(lst):
    vals = [v for v in lst if v is not None]
    if len(vals) < 2: return None
    for i in range(1, len(vals)):
        if vals[i] < vals[i-1]: return False
    return True

# ── 列インデックス検索（部分一致） ───────────────────
def find_col(hdr, keywords):
    for i, h in enumerate(hdr):
        if any(k in h for k in keywords):
            return i
    return None

# ── irbank /results ──────────────────────────────────
def fetch_irbank_results(code):
    """売上(百万), 営利率(%), EPS(円), 自己資本比率(%), 営業CF(百万) を10年分"""
    try:
        r = requests.get(f'https://irbank.net/{code}/results',
                         headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.content, 'html.parser')
        tables = soup.find_all('table')
        if not tables: return None

        # ── Table0: 売上, 営利率, EPS ──
        t0 = tables[0]
        rows0 = t0.find_all('tr')
        hdr0 = [c.get_text(strip=True) for c in rows0[0].find_all(['th','td'])]

        sales_i  = find_col(hdr0, ['売上', '営業収益', '経常収益', '純営業収益', '収益',
                                    '正味収入保険料', '営業総収入', '営業収入', '完成工事高'])
        eps_i    = find_col(hdr0, ['EPS'])
        oprate_i = find_col(hdr0, ['営利率'])

        main = {}  # label -> {sales, eps, op_margin}
        for row in rows0[1:]:
            cells = [c.get_text(strip=True) for c in row.find_all(['th','td'])]
            if not cells or not cells[0]: continue
            yr_m = re.search(r'(\d{4})', cells[0])
            if not yr_m or int(yr_m.group(1)) > TODAY_YR: continue
            lb = yr_label(cells[0]) or yr_m.group(1)
            main[lb] = {
                'sales':     parse_oku(cells[sales_i])   if sales_i  is not None and sales_i  < len(cells) else None,
                'eps':       parse_num(cells[eps_i])     if eps_i    is not None and eps_i    < len(cells) else None,
                'op_margin': parse_num(cells[oprate_i])  if oprate_i is not None and oprate_i < len(cells) else None,
            }

        # ── Table1: 自己資本比率 ──
        eq_data = {}
        if len(tables) > 1:
            t1 = tables[1]
            rows1 = t1.find_all('tr')
            hdr1 = [c.get_text(strip=True) for c in rows1[0].find_all(['th','td'])]
            eq_i = find_col(hdr1, ['自己資本比率', '株主資本比率'])
            if eq_i is not None:
                for row in rows1[1:]:
                    cells = [c.get_text(strip=True) for c in row.find_all(['th','td'])]
                    if not cells or not cells[0]: continue
                    yr_m = re.search(r'(\d{4})', cells[0])
                    if not yr_m or int(yr_m.group(1)) > TODAY_YR: continue
                    lb = yr_label(cells[0]) or yr_m.group(1)
                    eq_data[lb] = parse_num(cells[eq_i]) if eq_i < len(cells) else None

        # ── Table2: 営業CF + 現金等 (億円→百万円) ──
        cf_data = {}
        cf_cash_data = {}  # 現金等（BSで取れない場合のフォールバック）
        if len(tables) > 2:
            t2 = tables[2]
            rows2 = t2.find_all('tr')
            hdr2 = [c.get_text(strip=True) for c in rows2[0].find_all(['th','td'])]
            ocf_i   = find_col(hdr2, ['営業CF'])
            cash_i2 = find_col(hdr2, ['現金等', '現金及び現金同等物'])
            for row in rows2[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(['th','td'])]
                if not cells or not cells[0]: continue
                yr_m = re.search(r'(\d{4})', cells[0])
                if not yr_m or int(yr_m.group(1)) > TODAY_YR: continue
                lb = yr_label(cells[0]) or yr_m.group(1)
                if ocf_i is not None:
                    val = parse_oku(cells[ocf_i]) if ocf_i < len(cells) else None
                    if val is not None:
                        cf_data[lb] = val
                if cash_i2 is not None:
                    val2 = parse_oku(cells[cash_i2]) if cash_i2 < len(cells) else None
                    if val2 is not None:
                        cf_cash_data[lb] = val2

        # 最大10年
        sy = sorted(main.keys())[-10:]
        ey = sorted(eq_data.keys())[-10:]
        cy = sorted(cf_data.keys())[-10:]
        ccy = sorted(cf_cash_data.keys())[-10:]
        return {
            'years':     sy,
            'sales':     [main[y]['sales']     for y in sy],
            'eps':       [main[y]['eps']       for y in sy],
            'op_margin': [main[y]['op_margin'] for y in sy],
            'eq_years':  ey,
            'equity':    [eq_data[y]           for y in ey],
            'cf_years':  cy,
            'ocf':       [cf_data[y]           for y in cy],
            'cf_cash_years': ccy,
            'cf_cash':   [cf_cash_data[y]      for y in ccy],
        }
    except: return None

# ── irbank /bs 現金及び預金 ───────────────────────────
def fetch_irbank_bs_cash(code, ref_years=None):
    """現金及び預金（irbank BS、10年）。ref_yearsで年次のみに絞る"""
    try:
        r = requests.get(f'https://irbank.net/{code}/bs',
                         headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.content, 'html.parser')
        tables = soup.find_all('table')
        if not tables: return None
        t = tables[0]
        rows = t.find_all('tr')
        if not rows: return None
        hdr = [c.get_text(strip=True) for c in rows[0].find_all(['th','td'])]
        cash_i = find_col(hdr, ['現金及び預金', '現金等', '現金・預金', '現金及び現金同等物'])
        if cash_i is None: return None
        cash_data = {}
        for row in rows[1:]:
            cells = [c.get_text(strip=True) for c in row.find_all(['th','td'])]
            if not cells or not cells[0]: continue
            yr_m = re.search(r'(\d{4})', cells[0])
            if not yr_m or int(yr_m.group(1)) > TODAY_YR: continue
            lb = yr_label(cells[0]) or yr_m.group(1)
            val = parse_num(cells[cash_i]) if cash_i < len(cells) else None
            if val is not None:
                cash_data[lb] = val
        if ref_years:
            cy = [y for y in sorted(cash_data.keys()) if y in set(ref_years)][-10:]
        else:
            cy = sorted(cash_data.keys())[-10:]
        return {'cash_years': cy, 'cash': [cash_data[y] for y in cy]}
    except: return None

# ── irbank 配当金履歴 ─────────────────────────────────
def fetch_irbank_dividend(code):
    try:
        r = requests.get(f'https://irbank.net/{code}/dividend',
                         headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.content, 'html.parser')
        for t in soup.find_all('table'):
            rows = t.find_all('tr')
            if not rows: continue
            hdr = [c.get_text(strip=True) for c in rows[0].find_all(['th','td'])]
            if '合計' not in hdr and '分割調整' not in hdr: continue
            col = hdr.index('分割調整') if '分割調整' in hdr else hdr.index('合計')
            data = []
            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(['th','td'])]
                if len(cells) <= col: continue
                val = parse_num(cells[col])
                if val is None or val <= 0: continue
                yr_m = re.search(r'(\d{4})年', cells[0])
                if not yr_m: continue
                yr = int(yr_m.group(1))
                if yr > TODAY_YR: continue
                lb = yr_label(cells[0]) or str(yr)
                data.append((yr, lb, val))
            data.sort(key=lambda x: x[0])
            return [(lb, v) for _, lb, v in data]
        return []
    except: return []

# ── エントリ構築 ──────────────────────────────────────
def build_entry(code, name, yld, res, bs_cash, div):
    if not res or not res.get('years'): return None

    years      = res['years']
    sales      = res['sales']
    op_margin  = res['op_margin']
    eps        = res['eps']
    eq_years   = res.get('eq_years', [])
    equity     = res.get('equity', [])

    cf_years   = res.get('cf_years', [])
    ocf        = res.get('ocf', [])
    # BSキャッシュが空の場合はCFテーブルの現金等をフォールバック
    if bs_cash and bs_cash.get('cash'):
        cash_years = bs_cash['cash_years']
        cash       = bs_cash['cash']
    elif res.get('cf_cash'):
        cash_years = res.get('cf_cash_years', [])
        cash       = res.get('cf_cash', [])
    else:
        cash_years = []
        cash       = []

    div_years  = [d[0] for d in div]
    div_vals   = [d[1] for d in div]


    # 配当性向: EPS と div を年ラベルでマッチ
    eps_dict = dict(zip(years, eps))
    payout_years, payout_vals = [], []
    for lb, dv in zip(div_years, div_vals):
        e = eps_dict.get(lb)
        if e is not None and e > 0 and dv is not None:
            payout_years.append(lb)
            payout_vals.append(round(dv / e * 100, 1))

    # 判定
    ocf_nonnull = [v for v in ocf if v is not None]
    pass_ocf    = (not any(v <= 0 for v in ocf_nonnull)) if ocf_nonnull else None
    om_latest   = next((v for v in reversed(op_margin) if v is not None), None)
    pass_om     = (om_latest >= 10) if om_latest is not None else None
    eq_latest   = next((v for v in reversed(equity)    if v is not None), None)
    pass_eq     = (eq_latest >= 40) if eq_latest is not None else None
    nc          = no_cut(div_vals) if len(div_vals) >= 2 else None
    pass_div    = True if nc else (False if nc is False else None)
    po_latest   = next((v for v in reversed(payout_vals) if v is not None), None)
    pass_po     = (30 <= po_latest <= 50) if po_latest is not None else None

    verdicts = {
        'sales':     check_trend(sales),
        'op_margin': pass_om,
        'eps':       check_trend(eps, min_ratio=1.0),
        'ocf':       pass_ocf,
        'div':       pass_div,
        'payout':    pass_po,
        'equity':    pass_eq,
        'cash':      check_trend(cash),
    }
    score = sum(1 for v in verdicts.values() if v is True)

    return {
        'name': name, 'code': code, 'yield': round(yld, 2),
        'score': score, 'fetch_date': str(datetime.date.today()),
        'years':       years,     'sales':    sales,
        'op_margin':   op_margin, 'eps':      eps,
        'cf_years':    cf_years,  'ocf':      ocf,
        'cash_years':  cash_years,'cash':     cash,
        'eq_years':    eq_years,  'equity':   equity,
        'div_years':   div_years, 'div':      div_vals,
        'payout_years':payout_years, 'payout': payout_vals,
        'verdicts':    verdicts,
    }

# ── 銘柄リスト読み込み ────────────────────────────────
print('銘柄リスト読み込み...')
wb = load_workbook(BASE_DIR / '5以上_銘柄リスト_v2.xlsx')
ws = wb['5以上判定一覧']
stocks = []
for r in range(5, ws.max_row+1):
    v = ws.cell(r, 2).value
    if not v or not str(v).strip(): continue
    if str(ws.cell(r, 1).value or '').startswith('━'): continue
    try: score = int(ws.cell(r, 5).value or 0)
    except: continue
    if score < 5: continue
    stocks.append({'code': str(v).strip(),
                   'name': str(ws.cell(r, 3).value or ''),
                   'yield': float(ws.cell(r, 4).value or 0)})
print(f'対象: {len(stocks)}銘柄')

# ── キャッシュ読み込み ────────────────────────────────
cache_path = BASE_DIR / 'chart_data_cache.json'
if os.path.exists(cache_path):
    with open(cache_path, encoding='utf-8') as f:
        all_data = json.load(f)
    # 旧形式(payout_years なし)は再取得
    stale = [k for k,v in all_data.items() if 'payout_years' not in v]
    for k in stale: del all_data[k]
    print(f'キャッシュ: {len(all_data)}銘柄（旧形式 {len(stale)}件を除去）')
else:
    all_data = {}

# ── データ取得 ────────────────────────────────────────
new_count = 0
for i, s in enumerate(stocks, 1):
    code = s['code']
    if code in all_data:
        print(f'  [{i:3d}/{len(stocks)}] {code} {s["name"]} (cache)')
        continue
    print(f'  [{i:3d}/{len(stocks)}] {code} {s["name"]}', end=' ', flush=True)

    res     = fetch_irbank_results(code);  time.sleep(0.6)
    bs_cash = fetch_irbank_bs_cash(code, ref_years=res.get('years',[]) if res else []);  time.sleep(0.6)
    div     = fetch_irbank_dividend(code); time.sleep(0.6)

    entry = build_entry(code, s['name'], s['yield'], res, bs_cash, div)
    if entry:
        all_data[code] = entry
        new_count += 1
        print(f'→ {entry["score"]}/8  ({len(entry["years"])}yr)')
    else:
        print('→ データなし')

    if new_count % 10 == 0:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(all_data, f, ensure_ascii=False)

with open(cache_path, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False)
print(f'\nキャッシュ保存: {len(all_data)}銘柄')

# ── 業種・決算発表日キャッシュ ───────────────────────
from collections import defaultdict
sector_cache_path = BASE_DIR / 'sector_cache.json'
earnings_date_cache_path = BASE_DIR / 'earnings_date_cache.json'

if os.path.exists(sector_cache_path):
    with open(sector_cache_path, encoding='utf-8') as f:
        sector_cache = json.load(f)
    print(f'業種キャッシュ: {len(sector_cache)}件')
else:
    sector_cache = {}

if os.path.exists(earnings_date_cache_path):
    with open(earnings_date_cache_path, encoding='utf-8') as f:
        earnings_date_cache = json.load(f)
    print(f'決算日キャッシュ: {len(earnings_date_cache)}件')
else:
    earnings_date_cache = {}

info_new = 0
all_codes = list(all_data.keys())
for i, code in enumerate(all_codes, 1):
    need_sector = code not in sector_cache
    need_date   = code not in earnings_date_cache
    if not need_sector and not need_date:
        continue
    sector, earnings_date = fetch_irbank_main(code)
    if need_sector:
        sector_cache[code] = sector or '不明'
    if need_date:
        earnings_date_cache[code] = earnings_date
    info_new += 1
    print(f'  情報取得 [{i}/{len(all_codes)}] {code} → {sector_cache.get(code,"不明")} / {earnings_date_cache.get(code,"－")}')
    time.sleep(0.5)
    if info_new % 30 == 0:
        with open(sector_cache_path, 'w', encoding='utf-8') as f:
            json.dump(sector_cache, f, ensure_ascii=False)
        with open(earnings_date_cache_path, 'w', encoding='utf-8') as f:
            json.dump(earnings_date_cache, f, ensure_ascii=False)

with open(sector_cache_path, 'w', encoding='utf-8') as f:
    json.dump(sector_cache, f, ensure_ascii=False)
with open(earnings_date_cache_path, 'w', encoding='utf-8') as f:
    json.dump(earnings_date_cache, f, ensure_ascii=False)
if info_new:
    print(f'情報取得完了: {info_new}件')
else:
    print('キャッシュ: 全件済み')

# ── 業種別 営業利益率 平均を計算 ──────────────────────
sector_om = defaultdict(lambda: defaultdict(list))
for code, d in all_data.items():
    sec = sector_cache.get(code, '不明')
    for y, v in zip(d.get('years', []), d.get('op_margin', [])):
        if v is not None:
            sector_om[sec][y].append(v)

sector_avg = {}
for sec, yd in sector_om.items():
    sector_avg[sec] = {y: round(sum(vs)/len(vs), 1) for y, vs in yd.items()}
print(f'業種別営業利益率平均: {len(sector_avg)}業種')

# ── HTML テンプレート ─────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>高配当株 8項目チャートビューア</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Yu Gothic','Meiryo',sans-serif;background:#EEF2F7;color:#1a1a2e}
header{background:linear-gradient(135deg,#1F3864,#2E75B6);color:#fff;padding:14px 24px;
       display:flex;align-items:center;gap:16px;flex-wrap:wrap}
header h1{font-size:1.2rem;font-weight:900}
.sb{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.sb input{padding:8px 14px;border:none;border-radius:8px;font-size:1rem;
          width:150px;outline:none;font-family:inherit}
.sb button{padding:8px 20px;background:#fff;color:#1F3864;border:none;
           border-radius:8px;cursor:pointer;font-weight:700;font-size:0.9rem}
.sb button:hover{background:#e0e8f4}
.ac{position:relative;display:inline-block}
.acl{position:absolute;top:100%;left:0;background:#fff;border:1px solid #BDD7EE;
     border-radius:8px;min-width:260px;max-height:220px;overflow-y:auto;
     z-index:999;box-shadow:0 4px 16px rgba(0,0,0,0.12);display:none}
.aci{padding:8px 14px;cursor:pointer;font-size:0.88rem;border-bottom:1px solid #f0f0f0;
     display:flex;justify-content:space-between;align-items:center}
.aci:hover,.aci.act{background:#EEF2F7}
.aci-code{font-weight:700;color:#1F3864;margin-right:8px;flex-shrink:0}
.aci-name{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.aci-sc{font-weight:700;font-size:0.82rem;margin-left:8px;flex-shrink:0}
.s8{color:#276221}.s7{color:#2E75B6}.s6{color:#7030A0}.s5{color:#843C0C}
#ib{background:#fff;padding:10px 24px;display:flex;align-items:center;
    gap:8px;flex-wrap:wrap;border-bottom:2px solid #BDD7EE;min-height:52px}
#it{font-size:1.1rem;font-weight:900;color:#888}
.badge{display:inline-flex;align-items:center;padding:3px 11px;border-radius:16px;
       font-size:0.78rem;font-weight:700;margin:2px}
.bp{background:#C6EFCE;color:#276221}
.bf{background:#FCE4D6;color:#843C0C}
.bn{background:#FFEB9C;color:#7d6608}
.sbadge{background:#1F3864;color:#fff;font-size:0.95rem;padding:4px 14px;
        border-radius:16px;font-weight:900}
#ilinks{display:flex;gap:8px;margin-left:4px}
.ext-link{display:inline-flex;align-items:center;gap:5px;padding:4px 12px;
          border-radius:8px;font-size:0.8rem;font-weight:700;text-decoration:none;
          border:1.5px solid;transition:opacity .15s}
.ext-link:hover{opacity:0.75}
.lk-ir{color:#e85d04;border-color:#e85d04;background:#fff8f4}
.lk-kb{color:#0369a1;border-color:#0369a1;background:#f0f9ff}
.lk-td{color:#7030A0;border-color:#7030A0;background:#f8f0ff}
.lk-mk{color:#e63946;border-color:#e63946;background:#fff0f1}
.grid{display:grid;grid-template-columns:repeat(3,1fr);
      gap:16px;padding:20px}
.card{background:#fff;border-radius:12px;box-shadow:0 2px 10px rgba(0,0,0,0.07);padding:18px}
.ch{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:3px}
.ct{font-size:0.93rem;font-weight:700;color:#1F3864}
.cn{font-size:0.72rem;color:#777;margin-bottom:10px}
.vd{width:32px;height:32px;border-radius:50%;display:flex;align-items:center;
    justify-content:center;font-weight:900;font-size:1.2rem;flex-shrink:0}
.vp{background:#C6EFCE;color:#276221}
.vf{background:#FCE4D6;color:#843C0C}
.vn{background:#FFEB9C;color:#7d6608}
.cw{position:relative;height:195px}
.rl{font-size:0.7rem;color:#999;margin-top:5px}
#ftbl th{background:#1F3864;color:#fff;padding:8px 12px;text-align:right;font-size:0.8rem;font-weight:700;white-space:nowrap}
#ftbl th:first-child{text-align:left}
#ftbl td{padding:7px 12px;text-align:right;border-bottom:1px solid #f0f0f0;white-space:nowrap}
#ftbl td:first-child{text-align:left;font-weight:700;color:#444}
#ftbl tr:hover td{background:#f5f8ff}
#ftbl tr.hl-row td{background:#EEF2F7}
.fc-up{color:#276221;font-weight:700}
.fc-dn{color:#843C0C;font-weight:700}
.fc-eq{color:#555}
</style>
</head>
<body>
<header>
  <h1>高配当株 8項目チャートビューア</h1>
  <div class="sb">
    <div class="ac">
      <input id="ci" type="text" placeholder="コード or 銘柄名" autocomplete="off">
      <div class="acl" id="al"></div>
    </div>
    <button onclick="go()">表示</button>
  </div>
  <span style="font-size:0.8rem;opacity:0.85">{total}銘柄収録（5/8以上）　①②③⑦⑧=10年 ④=3〜5年 ⑤⑥=10年+</span>
</header>
<div id="ib"><span id="it">銘柄コードまたは銘柄名を入力してください</span><div id="ilinks"></div><div id="ibg"></div></div>
<div class="grid" id="grid"></div>
<div id="ftbl-wrap" style="display:none;padding:20px">
  <div style="font-weight:900;font-size:0.95rem;color:#1F3864;margin-bottom:8px">📋 財務データ一覧</div>
  <div style="overflow-x:auto">
    <table id="ftbl" style="width:100%;border-collapse:collapse;font-size:0.85rem;background:#fff;border-radius:12px;box-shadow:0 2px 10px rgba(0,0,0,0.07);overflow:hidden"></table>
  </div>
</div>

<script>
const D = {data_json};
const SECTOR = {sector_json};
const SECTOR_AVG = {sector_avg_json};
const ED = {earnings_date_json};

// バッジ表示用（8項目全部）
const BADGE_ITEMS = [
  {k:'sales',    lb:'① 売上高'},
  {k:'op_margin',lb:'② 営業利益率'},
  {k:'eps',      lb:'③ EPS'},
  {k:'ocf',      lb:'④ 営業CF'},
  {k:'div',      lb:'⑤ 1株配当金'},
  {k:'payout',   lb:'⑥ 配当性向'},
  {k:'equity',   lb:'⑦ 自己資本比率'},
  {k:'cash',     lb:'⑧ 現金等残高'},
];

// チャート定義（6チャート）
// type: 'bar' = 棒グラフのみ, 'line' = 折れ線のみ, 'combo' = 左右2軸コンボ
const CHARTS = [
  {
    id: 0,
    title: '① 売上高 ＋ ② 営業利益率',
    type: 'combo',
    note: '売上高: 右肩上がり（前年比90%以上）／ 営利率: 10%以上',
    rl: '営利率基準 10%',
    left:  {k:'sales',    lb:'売上高（百万円）',  verdictKey:'sales',    barColor:'rgba(46,117,182,0.75)'},
    right: {k:'op_margin',lb:'営業利益率（%）',   verdictKey:'op_margin', refLine:{val:10, label:'基準 10%', color:'rgba(0,176,80,0.85)'}},
  },
  {
    id: 1,
    title: '③ EPS',
    type: 'bar',
    note: '右肩上がり（前年比90%以上）',
    rl: null,
    k:'eps', u:'円', verdictKey:'eps',
    getYears: d=>d.years, getVals: d=>d.eps,
  },
  {
    id: 2,
    title: '④ 営業CF',
    type: 'bar',
    note: '過去10年で一度でも赤字→リスト対象外',
    rl: '基準 0（一度でも赤字→×）',
    k:'ocf', u:'百万円', verdictKey:'ocf',
    refLine: {val:0, label:'基準 0', color:'rgba(231,76,60,0.7)'},
    getYears: d=>d.cf_years, getVals: d=>d.ocf,
  },
  {
    id: 3,
    title: '⑤ 1株配当金 ＋ ⑥ 配当性向',
    type: 'combo',
    note: '1株配当: 減配なし・右肩上がり／ 配当性向: 30〜50%',
    rl: '配当性向基準 30〜50%',
    left:  {k:'div',    lb:'1株配当金（円）',  verdictKey:'div',    barColor:'rgba(46,117,182,0.75)'},
    right: {k:'payout', lb:'配当性向（%）',    verdictKey:'payout', refBand:{lo:30, hi:50, loLabel:'下限30%', hiLabel:'上限50%'}},
    // 1株配当とpayoutはx軸が異なる可能性 → union でマージ
    mergeXAxis: true,
  },
  {
    id: 4,
    title: '⑦ 自己資本比率',
    type: 'line',
    note: '40%以上（直近期）',
    rl: '基準 40%',
    k:'equity', u:'%', verdictKey:'equity',
    refLine: {val:40, label:'基準 40%', color:'rgba(0,176,80,0.85)'},
    getYears: d=>d.eq_years, getVals: d=>d.equity,
  },
  {
    id: 5,
    title: '⑧ 現金等残高',
    type: 'bar',
    note: '右肩上がり（前年比90%以上）',
    rl: null,
    k:'cash', u:'百万円', verdictKey:'cash',
    getYears: d=>d.cash_years, getVals: d=>d.cash,
  },
];

function gj(k,d){return d.verdicts[k]!==undefined?d.verdicts[k]:null;}

// 棒グラフ用カラー（前年比較）
function barColors(vl){
  return vl.map((v,i)=>{
    if(v===null)return 'rgba(180,180,180,0.3)';
    let prev=null;
    for(let j=i-1;j>=0;j--){if(vl[j]!==null){prev=vl[j];break;}}
    if(prev===null)return 'rgba(46,117,182,0.75)';
    return v>=prev?'rgba(46,117,182,0.75)':'rgba(231,76,60,0.78)';
  });
}

// x軸 union マージ（div_years と payout_years を統合、欠損はnull埋め）
function mergeYears(yrsA, valsA, yrsB, valsB){
  const all=[...new Set([...yrsA,...yrsB])].sort();
  const mapA=Object.fromEntries(yrsA.map((y,i)=>[y,valsA[i]]));
  const mapB=Object.fromEntries(yrsB.map((y,i)=>[y,valsB[i]]));
  return {
    labels: all,
    dataA:  all.map(y=>mapA[y]!==undefined?mapA[y]:null),
    dataB:  all.map(y=>mapB[y]!==undefined?mapB[y]:null),
  };
}

const CI={};

function buildComboChart(canvasId, chartDef, d, sectorAvgLine=null){
  const cfg=chartDef;
  let labels, leftData, rightData;

  if(cfg.mergeXAxis){
    // div + payout: x軸をマージ
    const yrsL=d[cfg.left.k==='div'?'div_years':'years']||[];
    const vlsL=d[cfg.left.k]||[];
    const yrsR=d[cfg.right.k==='payout'?'payout_years':'years']||[];
    const vlsR=d[cfg.right.k]||[];
    const merged=mergeYears(yrsL,vlsL,yrsR,vlsR);
    labels=merged.labels; leftData=merged.dataA; rightData=merged.dataB;
  } else {
    // 同じx軸（years）
    labels=d.years||[];
    leftData=d[cfg.left.k]||[];
    rightData=d[cfg.right.k]||[];
  }

  // 左軸データが全部nullの場合は棒グラフなし（営業利益率のみ表示）
  const hasLeftData = leftData.some(v=>v!==null);
  const bc=hasLeftData?barColors(leftData):[];
  const datasets=hasLeftData?[
    {
      label: cfg.left.lb,
      data: leftData,
      type: 'bar',
      yAxisID: 'y',
      backgroundColor: bc,
      borderColor: bc.map(x=>x.replace('0.75','1').replace('0.3','0.5').replace('0.78','1')),
      borderWidth: 1.5,
      borderRadius: 4,
      order: 2,
    },
    {
      label: cfg.right.lb,
      data: rightData,
      type: 'line',
      yAxisID: 'y2',
      borderColor: 'rgba(231,76,60,0.9)',
      backgroundColor: 'transparent',
      borderWidth: 2,
      pointRadius: 3,
      pointBackgroundColor: 'rgba(231,76,60,0.9)',
      tension: 0.1,
      order: 1,
    },
  ]:[
    // 売上高データなし → 営業利益率のみ単軸で表示
    {
      label: cfg.right.lb,
      data: rightData,
      type: 'line',
      yAxisID: 'y',
      borderColor: 'rgba(231,76,60,0.9)',
      backgroundColor: 'rgba(231,76,60,0.1)',
      borderWidth: 2,
      pointRadius: 3,
      pointBackgroundColor: 'rgba(231,76,60,0.9)',
      tension: 0.1,
      fill: false,
    },
  ];

  // 右軸の基準線・バンド
  if(cfg.right.refLine){
    datasets.push({
      label: cfg.right.refLine.label,
      data: labels.map(()=>cfg.right.refLine.val),
      type: 'line',
      yAxisID: 'y2',
      borderColor: cfg.right.refLine.color,
      borderDash: [5,3],
      pointRadius: 0,
      borderWidth: 1.5,
      backgroundColor: 'transparent',
      order: 0,
    });
  }
  if(sectorAvgLine){
    datasets.push({
      label: sectorAvgLine.label,
      data: labels.map(y=>sectorAvgLine.data[y]!==undefined?sectorAvgLine.data[y]:null),
      type: 'line',
      yAxisID: 'y2',
      borderColor: 'rgba(255,140,0,0.9)',
      borderDash: [6,3],
      pointRadius: 0,
      borderWidth: 2,
      backgroundColor: 'transparent',
      order: 0,
    });
  }
  if(cfg.right.refBand){
    datasets.push({
      label: cfg.right.refBand.loLabel,
      data: labels.map(()=>cfg.right.refBand.lo),
      type: 'line',
      yAxisID: 'y2',
      borderColor: 'rgba(0,176,80,0.7)',
      borderDash: [5,3],
      pointRadius: 0,
      borderWidth: 1.5,
      backgroundColor: 'transparent',
      order: 0,
    });
    datasets.push({
      label: cfg.right.refBand.hiLabel,
      data: labels.map(()=>cfg.right.refBand.hi),
      type: 'line',
      yAxisID: 'y2',
      borderColor: 'rgba(231,76,60,0.7)',
      borderDash: [5,3],
      pointRadius: 0,
      borderWidth: 1.5,
      backgroundColor: 'transparent',
      order: 0,
    });
  }

  const scales = hasLeftData ? {
    x:{ticks:{font:{size:9},maxRotation:50}},
    y:{position:'left', ticks:{font:{size:9}}, grace:'5%', title:{display:false}},
    y2:{position:'right', ticks:{font:{size:9}}, grace:'5%', grid:{drawOnChartArea:false}},
  } : {
    x:{ticks:{font:{size:9},maxRotation:50}},
    y:{ticks:{font:{size:9}}, grace:'5%'},
  };

  return new Chart(document.getElementById(canvasId).getContext('2d'),{
    type: 'bar',
    data: {labels, datasets},
    options:{
      responsive:true, maintainAspectRatio:false,
      plugins:{
        legend:{display:true,position:'top',labels:{font:{size:10},boxWidth:12}},
        tooltip:{callbacks:{label:x=>`${x.dataset.label}: ${x.raw!==null?Number(x.raw).toLocaleString():'n/a'}`}},
      },
      scales,
    },
  });
}

function buildSingleChart(canvasId, chartDef, d){
  const cfg=chartDef;
  const yr=cfg.getYears(d);
  const vl=cfg.getVals(d);
  const isLine=cfg.type==='line';

  const bc=isLine?null:barColors(vl);
  const datasets=isLine?[
    {
      label: cfg.u,
      data: vl,
      type: 'line',
      borderColor: 'rgba(46,117,182,0.9)',
      backgroundColor: 'rgba(46,117,182,0.12)',
      borderWidth: 2,
      pointRadius: 3,
      tension: 0.1,
      fill: false,
    }
  ]:[
    {
      label: cfg.u,
      data: vl,
      backgroundColor: bc,
      borderColor: bc.map(x=>x.replace('0.75','1').replace('0.3','0.5').replace('0.78','1')),
      borderWidth: 1.5,
      borderRadius: 4,
    }
  ];

  if(cfg.refLine){
    datasets.push({
      label: cfg.refLine.label,
      data: yr.map(()=>cfg.refLine.val),
      type: 'line',
      borderColor: cfg.refLine.color,
      borderDash: [5,3],
      pointRadius: 0,
      borderWidth: 2,
      backgroundColor: 'transparent',
    });
  }

  return new Chart(document.getElementById(canvasId).getContext('2d'),{
    type: isLine?'line':'bar',
    data: {labels:yr, datasets},
    options:{
      responsive:true, maintainAspectRatio:false,
      plugins:{
        legend:{display:datasets.length>1,position:'top',labels:{font:{size:10},boxWidth:12}},
        tooltip:{callbacks:{label:x=>`${x.dataset.label}: ${x.raw!==null?Number(x.raw).toLocaleString():'n/a'}`}},
      },
      scales:{
        x:{ticks:{font:{size:9},maxRotation:50}},
        y:{ticks:{font:{size:9}},grace:'5%'},
      },
    },
  });
}

function show(code){
  const d=D[code]; if(!d)return;
  const sc=Object.values(d.verdicts).filter(v=>v===true).length;
  const scl=sc>=8?'s8':sc>=7?'s7':sc>=6?'s6':'s5';
  const ed=ED[code];
  const edStr=ed?`　📅 次回決算: ${ed}`:'';
  document.getElementById('it').textContent=`【${code}】${d.name}　配当利回り: ${d.yield}%${edStr}`;
  document.getElementById('it').style.color='';
  document.getElementById('ilinks').innerHTML=
    `<a class="ext-link lk-ir" href="https://irbank.net/${code}" target="_blank">📊 IRバンク</a>`+
    `<a class="ext-link lk-kb" href="https://kabutan.jp/stock/finance?code=${code}" target="_blank">📈 株探</a>`+
    `<a class="ext-link lk-td" href="https://kabutan.jp/stock/news?code=${code}&categorycd=3" target="_blank">📋 決算短信</a>`+
    `<a class="ext-link lk-mk" href="https://minkabu.jp/stock/${code}/settlement" target="_blank">📋 みんかぶ</a>`;

  // バッジ（8項目全部）
  const sec=SECTOR[code]||'不明';
  let b=`<span class="badge sbadge ${scl}">${sc}/8</span><span class="badge" style="background:#FFF3CD;color:#856404">${sec}</span>`;
  BADGE_ITEMS.forEach(c=>{
    const v=gj(c.k,d);
    let lbl;
    if(c.k==='ocf'&&v===false) lbl='④ 営業CF ⚠リスト対象外';
    else lbl=`${c.lb} ${v===true?'○':v===false?'×':'－'}`;
    b+=`<span class="badge ${v===true?'bp':v===false?'bf':'bn'}">${lbl}</span>`;
  });
  document.getElementById('ibg').innerHTML=b;

  // 財務データ一覧テーブル
  buildFtbl(d);

  // グリッド（6チャート）
  const g=document.getElementById('grid'); g.innerHTML='';
  // 既存チャートを破棄
  Object.values(CI).forEach(c=>{try{c.destroy();}catch(e){}});

  CHARTS.forEach((cfg,i)=>{
    // カードのverdictは左軸キー（コンボの場合）or 単体キーで判定
    const vKey=cfg.type==='combo'?cfg.left.verdictKey:cfg.verdictKey;
    const vd=gj(vKey,d);
    const vc=vd===true?'vp':vd===false?'vf':'vn';
    const vs=vd===true?'○':vd===false?'×':'－';

    // 直近値の表示（コンボは左軸の直近値）
    let latestStr='';
    if(cfg.type==='combo'){
      const lvl=d[cfg.left.k]||[];
      const lt=lvl.filter(v=>v!==null).slice(-1)[0];
      const lu=cfg.left.lb;
      latestStr=lt!==undefined?`直近: ${Number(lt).toLocaleString()} ${lu}`:'データなし';
    } else {
      const vl=cfg.getVals(d);
      const lt=vl.filter(v=>v!==null).slice(-1)[0];
      latestStr=lt!==undefined?`直近: ${Number(lt).toLocaleString()} ${cfg.u}`:'データなし';
    }

    const el=document.createElement('div'); el.className='card';
    el.innerHTML=
      `<div class="ch"><div><div class="ct">${cfg.title}</div><div class="cn">${cfg.note} ／ ${latestStr}</div></div><div class="vd ${vc}">${vs}</div></div>`+
      `<div class="cw"><canvas id="c${i}"></canvas></div>`+
      (cfg.rl?`<div class="rl">${cfg.rl}</div>`:'');
    g.appendChild(el);

    if(cfg.type==='combo'){
      let sectorAvgLine=null;
      if(cfg.id===0){
        const sec=SECTOR[code]||'不明';
        const avgData=SECTOR_AVG[sec]||{};
        sectorAvgLine={label:`${sec}平均`, data:avgData};
      }
      CI[i]=buildComboChart('c'+i, cfg, d, sectorAvgLine);
    } else {
      CI[i]=buildSingleChart('c'+i, cfg, d);
    }
  });
}

function buildFtbl(d){
  const wrap=document.getElementById('ftbl-wrap');
  const tbl=document.getElementById('ftbl');
  // 年度を結合（sales/ocf/div/payout/equity/cash の各年を統合）
  const allYears=new Set([
    ...(d.years||[]),
    ...(d.cf_years||[]),
    ...(d.payout_years||[]),
    ...(d.eq_years||[]),
    ...(d.cash_years||[])
  ]);
  const yrList=[...allYears].sort().reverse();
  if(!yrList.length){wrap.style.display='none';return;}

  // 年→値のマップ作成
  function mkMap(yrs,vals){const m={};(yrs||[]).forEach((y,i)=>{m[y]=vals[i];});return m;}
  const mSales=mkMap(d.years,d.sales);
  const mOm=mkMap(d.years,d.op_margin);
  const mEps=mkMap(d.years,d.eps);
  const mOcf=mkMap(d.cf_years,d.ocf);
  const mDiv=mkMap(d.div_years,d.div);
  const mPo=mkMap(d.payout_years,d.payout);
  const mEq=mkMap(d.eq_years,d.equity);
  const mCash=mkMap(d.cash_years,d.cash);

  function fc(val,prev,fmt){
    if(val==null)return '<td>－</td>';
    const s=fmt(val);
    if(prev==null)return `<td class="fc-eq">${s}</td>`;
    if(val>prev)return `<td class="fc-up">${s}</td>`;
    if(val<prev)return `<td class="fc-dn">${s}</td>`;
    return `<td class="fc-eq">${s}</td>`;
  }
  function fmtN(v){return v!=null?Number(v).toLocaleString():'－';}
  function fmtP(v){return v!=null?v.toFixed(1)+'%':'－';}
  function fmtDiv(v){return v!=null?v+'円':'－';}

  let html='<thead><tr>'
    +'<th>年度</th><th>売上高（百万）</th><th>営業利益率</th><th>EPS</th>'
    +'<th>営業CF（百万）</th><th>1株配当</th><th>配当性向</th><th>自己資本比率</th><th>現金等（百万）</th>'
    +'</tr></thead><tbody>';

  yrList.forEach((yr,idx)=>{
    const nextYr=yrList[idx+1]; // 1つ古い年
    const rowCls=idx%2===1?' class="hl-row"':'';
    const label=yr.replace('/','年').replace(/\/.*/,'');
    html+=`<tr${rowCls}><td>${label}</td>`
      +fc(mSales[yr],mSales[nextYr],v=>Number(v).toLocaleString())
      +fc(mOm[yr],mOm[nextYr],v=>v.toFixed(1)+'%')
      +fc(mEps[yr],mEps[nextYr],v=>v.toFixed(2)+'円')
      +fc(mOcf[yr],mOcf[nextYr],v=>Number(v).toLocaleString())
      +fc(mDiv[yr],mDiv[nextYr],v=>v+'円')
      +fc(mPo[yr],mPo[nextYr],v=>v.toFixed(2)+'%')
      +fc(mEq[yr],mEq[nextYr],v=>v.toFixed(1)+'%')
      +fc(mCash[yr],mCash[nextYr],v=>Number(v).toLocaleString())
      +'</tr>';
  });
  html+='</tbody>';
  tbl.innerHTML=html;
  wrap.style.display='block';
}

const ci2=document.getElementById('ci'),al=document.getElementById('al');
let aa=-1;
const SL=Object.entries(D).map(([k,v])=>({code:k,name:v.name,score:v.score,yield:v.yield})).sort((a,b)=>b.score-a.score||b.yield-a.yield);
function sac(q){
  if(!q){al.style.display='none';return;}
  const m=SL.filter(s=>s.code.startsWith(q)||s.name.includes(q)).slice(0,15);
  if(!m.length){al.style.display='none';return;}
  al.innerHTML=m.map(s=>{const sc=s.score>=8?'s8':s.score>=7?'s7':s.score>=6?'s6':'s5';return`<div class="aci" data-code="${s.code}"><span class="aci-code">${s.code}</span><span class="aci-name">${s.name}</span><span class="aci-sc ${sc}">${s.score}/8</span></div>`;}).join('');
  al.style.display='block';aa=-1;
  al.querySelectorAll('.aci').forEach(e=>e.addEventListener('click',()=>{ci2.value=e.dataset.code;al.style.display='none';show(e.dataset.code);}));
}
ci2.addEventListener('input',()=>sac(ci2.value.trim()));
ci2.addEventListener('keydown',e=>{
  const it=al.querySelectorAll('.aci');
  if(e.key==='ArrowDown'){aa=Math.min(aa+1,it.length-1);it.forEach((x,i)=>x.classList.toggle('act',i===aa));}
  else if(e.key==='ArrowUp'){aa=Math.max(aa-1,-1);it.forEach((x,i)=>x.classList.toggle('act',i===aa));}
  else if(e.key==='Enter'){if(aa>=0&&it[aa]){ci2.value=it[aa].dataset.code;al.style.display='none';show(ci2.value);}else go();}
  else if(e.key==='Escape')al.style.display='none';
});
document.addEventListener('click',e=>{if(!e.target.closest('.ac'))al.style.display='none';});
function go(){const v=ci2.value.trim();if(!v)return;if(D[v]){show(v);al.style.display='none';return;}const h=SL.find(s=>s.name.includes(v));if(h){ci2.value=h.code;show(h.code);}else alert('「'+v+'」は収録されていません');}
</script>
</body>
</html>"""

out = BASE_DIR / 'chart_viewer.html'
html = HTML.replace('{data_json}', json.dumps(all_data, ensure_ascii=False)) \
           .replace('{total}', str(len(all_data))) \
           .replace('{sector_json}', json.dumps(sector_cache, ensure_ascii=False)) \
           .replace('{sector_avg_json}', json.dumps(sector_avg, ensure_ascii=False)) \
           .replace('{earnings_date_json}', json.dumps(earnings_date_cache, ensure_ascii=False))
with open(out, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'HTML生成: {out}  ({len(all_data)}銘柄)')
webbrowser.open('file:///' + str(out).replace('\\','/'))
